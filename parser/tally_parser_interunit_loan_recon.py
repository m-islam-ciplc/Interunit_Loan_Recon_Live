# tally_parser_interunit_loan_recon.py

import re
import pandas as pd
from openpyxl import load_workbook
from calendar import month_name
from typing import Tuple, Optional
import datetime
import itertools

def extract_statement_period(metadata: pd.DataFrame) -> Tuple[Tuple[str, str], str, Optional[int]]:
    period_pattern = re.compile(r'(\d{1,2}-[A-Za-z]{3}-\d{4})\s*to\s*(\d{1,2}-[A-Za-z]{3}-\d{4})')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = period_pattern.search(cell)
        if match:
            return (match.group(1), match.group(2)), cell, i
    return ("", ""), "", None

def extract_company_name(metadata: pd.DataFrame) -> Tuple[str, str, int]:
    """Extract the current company name from the ledger file metadata"""
    # Try multiple patterns to find the company name
    patterns = [
        r'Unit\s*:?[\s)]*([^)]+)',  # Unit: CompanyName
        r'([A-Za-z\s&.()/-]+)\s+Unit',  # CompanyName Unit
        r'([A-Za-z\s&.()/-]+)\s+Account',  # CompanyName Account
        r'([A-Za-z\s&.()/-]+)\s+Ledger',  # CompanyName Ledger
    ]
    
    for i, row in metadata.iterrows():
        cell = str(row[0])
        for pattern in patterns:
            match = re.search(pattern, cell, re.IGNORECASE)
            if match:
                company_name = match.group(1).strip()
                # Clean up common suffixes
                company_name = re.sub(r'\s*[Uu]nit\.?\s*$', '', company_name).strip()
                company_name = re.sub(r'\s*[Aa]ccount\.?\s*$', '', company_name).strip()
                company_name = re.sub(r'\s*[Ll]edger\.?\s*$', '', company_name).strip()
                # Debug print to see what's being extracted
                # print(f"DEBUG: Extracted company name: '{company_name}' from cell: '{cell}'")
                return company_name, cell, i
    
    # Fallback: use first non-empty cell
    for i, row in metadata.iterrows():
        cell = str(row[0]).strip()
        if cell and cell not in ['', 'None', 'nan']:
            # print(f"DEBUG: Using fallback company name: '{cell}' from cell: '{cell}'")
            return cell, cell, i
    
    return "Unknown Company", "", 0

def extract_counterparty(metadata: pd.DataFrame) -> Tuple[str, str, int]:
    """Extract the counterparty company name from interunit loan account"""
    pattern = re.compile(r'Inter\s*Unit\s+Loan\s+A/C-(\w+)', re.IGNORECASE)
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = pattern.search(cell)
        if match:
            counterparty = match.group(1).strip()
            # Handle known company name variations
            if counterparty == 'Geo':
                counterparty = 'GeoTex'
            return counterparty, cell, i
    return "", "", None

def clean(val) -> str:
    cleaned = str(val).strip() if val is not None else ""
    # Remove _x000D_ characters (Carriage Return)
    cleaned = cleaned.replace('_x000D_', ' ')
    # Remove actual carriage return and line feed characters
    cleaned = cleaned.replace('\r', ' ')
    cleaned = cleaned.replace('\n', ' ')
    # Remove multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()

def deduplicate_row(row, dup_map):
    res = row[:]
    for val, idxs in dup_map.items():
        found = False
        for i in idxs:
            if clean(res[i]) == val:
                if found:
                    res[i] = ""
                else:
                    found = True
    return res

def parse_tally_file(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    header_keywords = {"Date", "Particulars", "Vch Type", "Vch No.", "Debit", "Credit"}
    header_row_idx = next((i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                           if header_keywords.issubset({clean(c) for c in r})), None)
    if not header_row_idx:
        wb.close()
        raise ValueError("Header row not found.")

    metadata_rows = []
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx-1, values_only=True):
        metadata_rows.append([clean(c) for c in row])
    metadata = pd.DataFrame(metadata_rows)

    (period_start, period_end), _, period_row = extract_statement_period(metadata)
    current_company, _, company_row = extract_company_name(metadata)
    counterparty, _, counterparty_row = extract_counterparty(metadata)

    # ledger_date = ""
    # ledger_year = ""
    # if period_start and period_end:
    #     try:
    #         first_date = pd.to_datetime(period_start, format="%d-%b-%Y")
    #         last_date = pd.to_datetime(period_end, format="%d-%b-%Y")
    #         if first_date.month == last_date.month:
    #             ledger_date = month_name[first_date.month]
    #         if first_date.year == last_date.year:
    #             ledger_year = str(first_date.year)
    #     except Exception:
    #         pass

    # Optimize merged cells processing - only process if there are merged cells
    if ws.merged_cells.ranges:
        for rng in list(ws.merged_cells.ranges):
            val = ws[rng.coord.split(":")[0]].value
            ws.unmerge_cells(str(rng))
            for row in ws[rng.coord]:
                for cell in row:
                    cell.value = val

    headers = [clean(c.value) if c.value else f"Unnamed_{i+1}" for i, c in enumerate(ws[header_row_idx])]

    headers = ["dr_cr" if h == "Particulars" and i == headers.index("Particulars") else h for i, h in enumerate(headers)]
    particulars_index = headers.index("dr_cr") + 1
    if particulars_index < len(headers):
        headers[particulars_index] = "Particulars"

    num_cols = len(headers)

    # Optimize row processing with batch operations
    collapsed_rows = []
    entered_by_list = []
    current_row = None
    last_entered_by = ""
    
    # Pre-compile regex for better performance
    entered_by_pattern = re.compile(r"entered by\s*:\s*(.*)", re.IGNORECASE)
    
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cleaned = [clean(c) for c in row][:num_cols] + ["" for _ in range(num_cols - len(row))]
        entered_by_found = False
        
        # Optimize entered_by detection
        for idx, cell in enumerate(cleaned):
            if "entered by :" in cell.lower():
                for next_idx in range(idx + 1, len(cleaned)):
                    if cleaned[next_idx]:
                        last_entered_by = cleaned[next_idx]
                        break
                else:
                    match = entered_by_pattern.search(cell)
                    if match:
                        last_entered_by = match.group(1).strip()
                entered_by_found = True
                break
                
        if entered_by_found:
            continue
            
        # Optimize row merging logic
        should_merge = (
            (not cleaned[headers.index("Date")] if "Date" in headers else True)
            and (not cleaned[headers.index("dr_cr")] if "dr_cr" in headers else True)
            and (cleaned[headers.index("Particulars")] if "Particulars" in headers else False)
            and current_row is not None
        )
        
        if should_merge:
            idx = headers.index("Particulars")
            current_row[idx] = (current_row[idx] + " " + cleaned[idx]).strip()
        else:
            if current_row is not None:
                collapsed_rows.append(current_row)
                entered_by_list.append(last_entered_by)
                last_entered_by = ""
            current_row = cleaned
            
    if current_row is not None:
        collapsed_rows.append(current_row)
        entered_by_list.append(last_entered_by)

    wb.close()
    dedup_map = {v: idxs for v, idxs in pd.Series(collapsed_rows[0]).groupby(
        lambda x: x).groups.items() if len(idxs) > 1}
    data_rows = [deduplicate_row(row, dedup_map) for row in collapsed_rows]

    if all(clean(v).replace('.', '', 1).replace(',', '', 1).isdigit() or clean(v) == "" for v in data_rows[-1]):
        data_rows.pop(-1)
        entered_by_list.pop(-1)

    df = pd.DataFrame(data_rows, columns=headers).dropna(axis=1, how='all')
    df = df.loc[:, (df != '').any(axis=0)]
    df = df.loc[:, ~df.columns.str.match(r'Unnamed_\d+')]

    df['entered_by'] = entered_by_list

    # Add role column based on Debit/Credit
    def determine_role(row):
        debit_val = row.get('Debit', 0)
        credit_val = row.get('Credit', 0)
        
        # Convert to float, handle empty/None values
        try:
            debit_float = float(debit_val) if debit_val and str(debit_val).strip() != '' else 0
            credit_float = float(credit_val) if credit_val and str(credit_val).strip() != '' else 0
        except (ValueError, TypeError):
            return None
            
        if debit_float > 0:
            return 'Lender'
        elif credit_float > 0:
            return 'Borrower'
        else:
            return None
    
    df['role'] = df.apply(determine_role, axis=1)
    
    # Add lender and borrower columns based on role for each transaction
    def assign_lender_borrower(row):
        role = row.get('role')
        if role == 'Lender':
            # Current company is lending (Debit > 0), so they are the lender
            return current_company, counterparty
        elif role == 'Borrower':
            # Current company is borrowing (Credit > 0), so counterparty is the lender
            return counterparty, current_company
        else:
            return None, None
    
    # Apply the function to get lender and borrower for each row
    lender_borrower_pairs = df.apply(assign_lender_borrower, axis=1)
    df['lender'] = [pair[0] for pair in lender_borrower_pairs]
    df['borrower'] = [pair[1] for pair in lender_borrower_pairs]
    
    # Debug print to see what's being assigned
    # print(f"DEBUG: Current company: '{current_company}'")
    # print(f"DEBUG: Counterparty: '{counterparty}'")
    # print(f"DEBUG: Sample lender values: {df['lender'].head().tolist()}")
    # print(f"DEBUG: Sample borrower values: {df['borrower'].head().tolist()}")

    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(
            df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    if "Particulars" in df.columns:
        df = df[df["Particulars"].str.strip().str.lower() != "opening balance"]
        df = df[~df["Particulars"].str.strip().str.lower().str.startswith("closing balance")]

    def to_hex(val: str) -> str:
        try:
            return hex(int(float(val)))[2:]
        except Exception:
            return ""

    uids = []
    rownum = 1
    for i, row in df.iterrows():
        date_val = row.get("Date", "")
        credit_val = row.get("Credit", "")
        debit_val = row.get("Debit", "")
        balance_val = credit_val if (pd.notna(credit_val) and str(credit_val).strip() != "") else debit_val
        if pd.notna(date_val) and date_val != "":
            date_str = str(date_val).replace("-", "")
            hexdate = to_hex(date_str)
            try:
                hexbal = to_hex(round(float(str(balance_val).replace(",", ""))))
            except Exception:
                hexbal = ""
            # Add company prefix to make uid unique across files
            uid = f"{current_company}_{hexdate}_{hexbal}_{rownum:06d}"
            uids.append(uid)
            rownum += 1
        else:
            uids.append("")
    df["uid"] = uids
    
    # Add input_date column
    from datetime import datetime
    df['input_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Remove original_filename completely - we use pair_id for tracking
    
    cols = ["uid", "lender", "borrower", "statement_month", "statement_year", "role"] + \
        [c for c in df.columns if c not in ["uid", "lender", "borrower", "statement_month", "statement_year", "role"]]

    # df["statement_month"] = ledger_date
    # df["statement_year"] = ledger_year
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    from calendar import month_name
    df["statement_month"] = df["Date"].dt.month.apply(
        lambda x: month_name[int(x)] if pd.notnull(x) else ""
    )
    df["statement_year"] = df["Date"].dt.year.astype("Int64")


    # Validate that all rows have valid statement_month and statement_year
    if df["statement_month"].eq("").any() or df["statement_year"].isna().any():
        raise ValueError("Parsing failed: Some rows have missing or invalid Date values, so statement_month or statement_year could not be extracted.")

    df = df[cols]

    if "Debit" in df.columns:
        df["Debit"] = df["Debit"].apply(lambda x: None if str(x).strip() == '' else x)
    if "Credit" in df.columns:
        df["Credit"] = df["Credit"].apply(lambda x: None if str(x).strip() == '' else x)

    new_column_names = {
        "Date": "Date",
        "Particulars": "Particulars",
        "Vch Type": "Vch_Type",
        "Vch No.": "Vch_No",
        "Debit": "Debit",
        "Credit": "Credit",
    }
    df = df.rename(columns=new_column_names)
    return df



if __name__ == "__main__":
    # Set these for IDE/Run Code button usage
    # input_file = "Input_Files/Interunit Steel.xlsx"
    # sheet_name = "Sheet7"

    input_file = "Input_Files/Interunit GeoTex.xlsx"
    sheet_name = "Sheet8"



    import sys
    if len(sys.argv) == 3:
        input_file = sys.argv[1]
        sheet_name = sys.argv[2]

    try:
        df = parse_tally_file(input_file, sheet_name)
        output_file = f"Parsed_{input_file.split('/')[-1]}"
        df.to_excel(output_file, index=False)
        # print(f"Successfully parsed {input_file} and saved as {output_file}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
