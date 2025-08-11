"""
ExportService - Handles all Excel export operations.
This service extracts 400+ lines of complex Excel generation logic from app.py routes.
"""
import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional, List
from flask import send_from_directory, jsonify
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core import database


class ExportService:
    """Centralizes all Excel export operations with consistent formatting."""
    
    def __init__(self):
        self.export_folder = 'uploads'
        os.makedirs(self.export_folder, exist_ok=True)
    
    def format_amount(self, amount) -> str:
        """Format amount with 2 decimal places for Excel export.
        
        Args:
            amount: The amount value (can be string, float, or None)
            
        Returns:
            Formatted amount string with 2 decimal places (e.g., "1000.00")
        """
        if amount is None or amount == '' or amount == 0:
            return "0.00"
        
        try:
            # Convert to float and format with 2 decimal places
            formatted = "{:.2f}".format(float(amount))
            return formatted
        except (ValueError, TypeError):
            # If conversion fails, return as is
            return str(amount) if amount else "0.00"
    
    def export_matched_transactions(self, filters: Dict[str, Optional[str]]):
        """Export auto-matched transactions as Excel with proper formatting.
        
        Only includes high-confidence auto-matches:
        - PO, LC, LOAN_ID, FINAL_SETTLEMENT, and INTERUNIT_LOAN matches
        - These are automatically confirmed due to high confidence
        - Excludes manual matches that require verification"""
        try:
            # Get filtered data - only auto-matched records
            lender_company = filters.get('lender_company')
            borrower_company = filters.get('borrower_company')
            month = filters.get('month')
            year = filters.get('year')
            
            if lender_company and borrower_company:
                matches = database.get_auto_matched_data_by_companies(lender_company, borrower_company, month, year)
            else:
                matches = database.get_auto_matched_data()
            
            if not matches:
                return jsonify({'error': 'No auto-matched data found'}), 404
            
            # Process and format data
            export_rows = self._process_matched_data(matches)
            
            # Create Excel file with descriptive filename
            df = pd.DataFrame(export_rows)
            
            # Generate filename: Auto_Matched_Transactions_Company Pair_Statement Period
            filename_parts = ['Auto_Matched_Transactions']
            
            if lender_company and borrower_company:
                company_pair = f"{lender_company}-{borrower_company}"
                filename_parts.append(company_pair)
            
            if month and year:
                statement_period = f"{month}_{year}"
                filename_parts.append(statement_period)
            
            # Fallback to timestamp if no specific filters
            if len(filename_parts) == 1:
                filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
            
            export_filename = f"{'_'.join(filename_parts)}.xlsx"
            export_path = os.path.join(self.export_folder, export_filename)
            
            # Apply formatting
            self._save_formatted_excel(df, export_path, 'matched')
            
            return send_from_directory(self.export_folder, export_filename, as_attachment=True)
        
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    def export_unmatched_transactions(self, filters: Dict[str, Optional[str]]):
        """Export unmatched transactions as Excel."""
        try:
            # Get filtered data
            lender_company = filters.get('lender_company')
            borrower_company = filters.get('borrower_company')
            month = filters.get('month')
            year = filters.get('year')
            
            if lender_company and borrower_company:
                unmatched = database.get_unmatched_data_by_companies(lender_company, borrower_company, month, year)
            else:
                unmatched = database.get_unmatched_data()
            
            if not unmatched:
                return jsonify({'error': 'No unmatched data found'}), 404
            
            # Create DataFrame
            df = pd.DataFrame(unmatched)
            
            # Generate filename: Unmatched_Transactions_Company Pair_Statement Period
            filename_parts = ['Unmatched_Transactions']
            
            if lender_company and borrower_company:
                company_pair = f"{lender_company}-{borrower_company}"
                filename_parts.append(company_pair)
            
            if month and year:
                statement_period = f"{month}_{year}"
                filename_parts.append(statement_period)
            
            # Fallback to timestamp if no specific filters
            if len(filename_parts) == 1:
                filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
            
            export_filename = f"{'_'.join(filename_parts)}.xlsx"
            export_path = os.path.join(self.export_folder, export_filename)
            
            # Apply formatting
            self._save_formatted_excel(df, export_path, 'unmatched')
            
            return send_from_directory(self.export_folder, export_filename, as_attachment=True)
        
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    def export_filtered_data(self, filters: Dict[str, Optional[str]]):
        """Export filtered data to Excel."""
        try:
            # Build filters dict, removing None values
            clean_filters = {k: v for k, v in filters.items() if v is not None}
            
            data = database.get_data(clean_filters)
            if not data:
                return jsonify({'error': 'No data found'}), 404
            
            df = pd.DataFrame(data)
            
            # Generate filename: Tally_Data_Filters_Statement Period
            filename_parts = ['Tally_Data']
            
            if filters.get('lender') and filters.get('borrower'):
                company_pair = f"{filters['lender']}-{filters['borrower']}"
                filename_parts.append(company_pair)
            
            if filters.get('statement_month') and filters.get('statement_year'):
                statement_period = f"{filters['statement_month']}_{filters['statement_year']}"
                filename_parts.append(statement_period)
            
            # Fallback to timestamp if no specific filters
            if len(filename_parts) == 1:
                filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
            
            export_filename = f"{'_'.join(filename_parts)}.xlsx"
            export_path = os.path.join(self.export_folder, export_filename)
            
            # Simple export without special formatting
            df.to_excel(export_path, index=False)
            
            return send_from_directory(self.export_folder, export_filename, as_attachment=True)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    def _process_matched_data(self, matches: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process matched data for export format."""
        df = pd.DataFrame(matches)
        
        # Get dynamic lender/borrower names
        lender_name, borrower_name = self._get_company_names(df)
        
        export_rows = []
        for _, row in df.iterrows():
            # Determine lender and borrower records
            main_record_debit = float(row.get('Debit', 0) or 0)
            matched_record_debit = float(row.get('matched_Debit', 0) or 0)
            
            if main_record_debit > 0:
                # Main record is lender
                lender_record = self._extract_lender_data(row, 'main')
                borrower_record = self._extract_borrower_data(row, 'matched')
            elif matched_record_debit > 0:
                # Matched record is lender
                lender_record = self._extract_lender_data(row, 'matched')
                borrower_record = self._extract_borrower_data(row, 'main')
            else:
                # Default fallback
                lender_record = self._extract_lender_data(row, 'main')
                borrower_record = self._extract_borrower_data(row, 'matched')
            
            export_row = {
                **lender_record,
                **borrower_record,
                'Match_Method': row.get('match_method', ''),
                'Audit_Info': self._format_audit_info(row.get('audit_info', ''))
            }
            
            export_rows.append(export_row)
        
        return export_rows
    
    def _get_company_names(self, df: pd.DataFrame) -> tuple:
        """Extract dynamic company names from data."""
        lender_name = 'Lender'
        borrower_name = 'Borrower'
        
        if len(df) > 0:
            first_row = df.iloc[0]
            if first_row.get('Debit') and first_row.get('Debit') > 0:
                lender_name = first_row.get('lender', 'Lender')
                borrower_name = first_row.get('matched_lender', 'Borrower')
            elif first_row.get('matched_Debit') and first_row.get('matched_Debit') > 0:
                lender_name = first_row.get('matched_lender', 'Lender')
                borrower_name = first_row.get('lender', 'Borrower')
            else:
                if first_row.get('lender'):
                    lender_name = first_row['lender']
                if first_row.get('matched_lender') and first_row['matched_lender'] != first_row.get('lender'):
                    borrower_name = first_row['matched_lender']
                elif first_row.get('borrower'):
                    borrower_name = first_row['borrower']
        
        return lender_name, borrower_name
    
    def _extract_lender_data(self, row: pd.Series, record_type: str) -> Dict[str, Any]:
        """Extract lender data from row."""
        prefix = '' if record_type == 'main' else 'matched_'
        
        return {
            'Lender_UID': row.get(f'{prefix}uid' if prefix else 'uid', ''),
            'Lender_Date': row.get(f'{prefix}Date' if prefix else 'Date', ''),
            'Lender_Particulars': row.get(f'{prefix}particulars' if prefix else 'Particulars', ''),
            'Lender_Debit': self.format_amount(row.get(f'{prefix}Debit' if prefix else 'Debit', 0)),
            'Lender_Vch_Type': row.get(f'{prefix}Vch_Type' if prefix else 'Vch_Type', ''),
            'Lender_Role': 'Lender'
        }
    
    def _extract_borrower_data(self, row: pd.Series, record_type: str) -> Dict[str, Any]:
        """Extract borrower data from row."""
        prefix = '' if record_type == 'main' else 'matched_'
        
        return {
            'Borrower_UID': row.get(f'{prefix}uid' if prefix else 'uid', ''),
            'Borrower_Date': row.get(f'{prefix}Date' if prefix else 'Date', ''),
            'Borrower_Particulars': row.get(f'{prefix}particulars' if prefix else 'Particulars', ''),
            'Borrower_Credit': self.format_amount(row.get(f'{prefix}Credit' if prefix else 'Credit', 0)),
            'Borrower_Vch_Type': row.get(f'{prefix}Vch_Type' if prefix else 'Vch_Type', ''),
            'Borrower_Role': 'Borrower'
        }
    
    def _format_audit_info(self, audit_info: str) -> str:
        """Format audit information for export."""
        if not audit_info:
            return ''
        
        try:
            import json
            info = json.loads(audit_info) if isinstance(audit_info, str) else audit_info
            match_type = info.get('match_type', 'Unknown')
            
            formatted = f"Match Type: {match_type}\n"
            formatted += f"Method: {info.get('match_method', 'Unknown')}\n"
            
            # Format based on match type
            if match_type == 'INTERUNIT_LOAN':
                if info.get('lender_reference'):
                    formatted += f"Lender: {info['lender_reference']}\n"
                if info.get('borrower_reference'):
                    formatted += f"Borrower: {info['borrower_reference']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
            elif match_type == 'PO':
                if info.get('po_number'):
                    formatted += f"PO Number: {info['po_number']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
            elif match_type == 'LC':
                if info.get('lc_number'):
                    formatted += f"LC Number: {info['lc_number']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
            elif match_type == 'LOAN_ID':
                if info.get('loan_id'):
                    formatted += f"Loan ID: {info['loan_id']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
            elif match_type == 'SALARY':
                if info.get('person'):
                    formatted += f"Person: {info['person']}\n"
                if info.get('period'):
                    formatted += f"Period: {info['period']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
                if info.get('jaccard_score'):
                    formatted += f"Similarity: {(info['jaccard_score'] * 100):.1f}%\n"
            elif match_type == 'FINAL_SETTLEMENT':
                if info.get('person'):
                    formatted += f"Person: {info['person']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
                if info.get('match_reason'):
                    formatted += f"Match Reason: {info['match_reason']}\n"
            elif match_type == 'COMMON_TEXT':
                if info.get('common_text'):
                    formatted += f"Matched Text: {info['common_text']}\n"
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
                if info.get('jaccard_score'):
                    formatted += f"Similarity: {(info['jaccard_score'] * 100):.1f}%\n"
            else:
                # Generic fallback for other match types
                if info.get('lender_amount'):
                    formatted += f"Lender Amount: {info['lender_amount']}\n"
                if info.get('borrower_amount'):
                    formatted += f"Borrower Amount: {info['borrower_amount']}\n"
            
            return formatted.strip()
        except:
            return str(audit_info)
    
    def _save_formatted_excel(self, df: pd.DataFrame, export_path: str, export_type: str):
        """Save DataFrame with Excel formatting."""
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{export_type.title()} Transactions"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._apply_excel_formatting(ws, export_type)
        
        # Save workbook
        wb.save(export_path)
    
    def _apply_excel_formatting(self, worksheet, export_type: str):
        """Apply consistent Excel formatting."""
        # Header formatting
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Define text wrapping columns and their properties
        text_wrap_columns = {
            'Lender_Particulars': {'width': 40, 'height': 60},
            'Borrower_Particulars': {'width': 40, 'height': 60},
            'Audit_Info': {'width': 35, 'height': 80}
        }
        
        # Apply text wrapping and formatting for specific columns
        for col_idx, col in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(col[0].column)
            header_value = str(col[0].value) if col[0].value else ''
            
            # Check if this column needs text wrapping
            if header_value in text_wrap_columns:
                # Set column width for text wrap columns
                worksheet.column_dimensions[column_letter].width = text_wrap_columns[header_value]['width']
                
                # Apply text wrapping to all cells in this column (including header)
                for cell in col:
                    # Create text wrap alignment
                    wrap_alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
                    cell.alignment = wrap_alignment
                    
                    # Set row height for better text wrapping display
                    if cell.row > 1:  # Skip header row for height adjustment
                        worksheet.row_dimensions[cell.row].height = text_wrap_columns[header_value]['height']
            else:
                # For non-text wrap columns, apply standard formatting
                # Auto-adjust column width based on content
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Apply standard alignment for non-text wrap columns
                for cell in col:
                    if cell.row > 1:  # Skip header row
                        cell.alignment = Alignment(horizontal='left', vertical='top')
        
        # Freeze header row
        worksheet.freeze_panes = "A2"